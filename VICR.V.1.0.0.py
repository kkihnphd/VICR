import cv2
import random
import os
from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import simpledialog, messagebox, filedialog, Canvas, Toplevel
from tkinter.ttk import Progressbar
from PIL import Image, ImageTk, ImageDraw, ImageFont
from pathlib import Path
import threading
import tkinter as tk
import numpy as np

# Initialize the main window
root = Tk()
root.title("Well Selector")
root.state('zoomed')  # Open in full-screen mode

# Variables to store the start and end points of the rectangle
rect_start = None
rect_end = None
current_rectangle = None
rectangles = []  # List to store the rectangles' coordinates and labels
media_rectangles = {}  # Dictionary to store rectangles for each media file
mode = None  # Can be 'video' or 'image'
existing_labels = set()
global cropped_videos_dir
cropped_videos_dir = None
existing_labels = set()
global cropped_images_dir
cropped_images_dir = None
stage_names = []
# Get the screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Initialize the canvas for drawing rectangles
canvas = Canvas(root, cursor="cross")
canvas.pack(fill=BOTH, expand=True)

# Define colors for rectangles
colors = ['red', 'green', 'blue', 'yellow', 'magenta', 'cyan']
color_index = 0

class Video:
    def __init__(self, label, cropped_path, original_file_name):
        self.label = label
        self.cropped_path = cropped_path
        self.original_file_name = original_file_name
        self.stage_times = {name: 'N/A' for name in stage_names}
    def set_stage_times(self, times):
        for stage, time in times.items():
            if stage in self.stage_times:
                self.stage_times[stage] = time if time else 'N/A'


def frame_generator(video_path):
    cap = cv2.VideoCapture(video_path)
    while cap.isOpened():
        ret, frame = cap.read()
        if ret:
            yield frame
        else:
            break
    cap.release()

# Function to get the first frame of the video
def get_first_frame(video_path):
    vid = cv2.VideoCapture(video_path)
    success, frame = vid.read()
    vid.release()
    if success:
        return cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    else:
        return None

# Function to draw rectangles on the canvas
def draw_rectangle(event):
    global rect_start, rect_end, current_rectangle, color_index
    if rect_start is None:
        rect_start = (event.x, event.y)
        current_rectangle = canvas.create_rectangle(
            rect_start[0], rect_start[1], event.x, event.y,
            outline=colors[color_index], width=2)
    else:
        rect_end = (event.x, event.y)
        canvas.coords(current_rectangle,
                      rect_start[0], rect_start[1], rect_end[0], rect_end[1])

def collect_stage_times(video, cap_window):
    global stage_names
    stage_selection = Toplevel(root)
    stage_selection.title("Input Notes for " + video.label)

    entries = {}
    for stage in stage_names:
        Label(stage_selection, text=f"Notes for {stage}:").pack()
        entry = Entry(stage_selection)
        entry.pack()
        entries[stage] = entry

    def on_submit():
        times = {stage: entry.get() for stage, entry in entries.items()}
        video.set_stage_times(times)
        stage_selection.destroy()
        cv2.destroyWindow(cap_window)  # Close the video window

    submit_button = Button(stage_selection, text="Submit", command=on_submit)
    submit_button.pack()

    stage_selection.wait_window()  # Wait for the window to be closed
def select_well_label():
    well_label = None

    def confirm_selection():
        nonlocal well_label
        letter = selected_letter.get()
        number = selected_number.get()
        combined_label = f"{letter}{number}"
        if combined_label not in existing_labels:
            well_label = combined_label
            existing_labels.add(combined_label)  # Add the new label to the set
            label_selection.destroy()
        else:
            messagebox.showwarning("Duplicate Label", "This label already exists. Please choose a different one.")

    label_selection = Toplevel(root)
    label_selection.title("Select Well Label")
    label_selection.grab_set()  # Makes the window modal

    # Dropdown for well letter
    selected_letter = StringVar(label_selection)
    selected_letter.set("A")  # default value
    letter_menu = OptionMenu(label_selection, selected_letter, "A", "B", "C", "D", "E", "F", "G", "H", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z")
    letter_menu.pack()

    # Dropdown for well number
    selected_number = StringVar(label_selection)
    selected_number.set("1")  # default value
    number_menu = OptionMenu(label_selection, selected_number, *[str(n) for n in range(1, 37)])
    number_menu.pack()

    # Confirm button
    confirm_button = Button(label_selection, text="Confirm", command=confirm_selection)
    confirm_button.pack()

    label_selection.wait_window()  # Wait for the window to be closed

    return well_label

# Function to finalize the rectangle
def finalize_rectangle(event):
    global rect_start, rect_end, media_rectangles, current_rectangle, color_index, current_media_index, mode, video_paths, image_paths, cropped_videos_dir, cropped_images_dir
    if mode == 'video':
        if not cropped_videos_dir:
            cropped_videos_dir = os.path.join(Path(video_paths[0]).parent, "Cropped_Videos")
            os.makedirs(cropped_videos_dir, exist_ok=True)
        current_media_path = video_paths[current_media_index]
    elif mode == 'image':
        if not cropped_images_dir:
            cropped_images_dir = os.path.join(Path(image_paths[0]).parent, f"{Path(image_paths[0]).stem}_Cropped_Images")
            os.makedirs(cropped_images_dir, exist_ok=True)
        current_media_path = image_paths[current_media_index]

    original_file_name = os.path.basename(current_media_path)
    if current_media_path not in media_rectangles:
        media_rectangles[current_media_path] = []

    label = select_well_label()
    if label:
        media_rectangles[current_media_path].append((rect_start, rect_end, label, original_file_name))
        if rect_start and rect_end:
            confirm = messagebox.askyesno("Confirm", "Is the selected area correct?")
            if confirm:
                if mode == 'video':
                    hidden_file_path = os.path.join(cropped_videos_dir, '.well_original_mapping.txt')
                if mode == 'image':
                    hidden_file_path = os.path.join(cropped_images_dir, '.well_original_mapping.txt')
                with open(hidden_file_path, 'a') as hidden_file:
                    hidden_file.write(f"{label},{current_media_path}\n")
                color_index = (color_index + 1) % len(colors)
            else:
                existing_labels.remove(label)
                canvas.delete(current_rectangle)
        rect_start, rect_end, current_rectangle = None, None, None


# Function to update the canvas with the first frame of the video
def update_canvas(media):
    global canvas, screen_width, screen_height, scaling_factor, mode, current_media_index, video_paths, image_paths
    # Define a margin to account for the window's title bar and other UI elements
    ui_margin = 150  # This value can be adjusted based on the actual UI layout

    # Adjust the available screen height to account for the margin
    available_height = screen_height - ui_margin

    if mode == 'video':
        # Video mode: assume 'media' is the first frame of the video
        frame = media
        height, width, channels = frame.shape
        # Get the filename of the current video
        filename = os.path.basename(video_paths[current_media_index])
    elif mode == 'image':
        # Image mode: assume 'media' is the path to the image file
        image = Image.open(media)
        width, height = image.size
        frame = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
        # Get the filename of the current image
        filename = os.path.basename(image_paths[current_media_index])

    # Calculate the scaling factor while maintaining the aspect ratio
    scaling_factor = min(screen_width / width, available_height / height, 1)

    # Apply the scaling factor
    resized_width = int(width * scaling_factor)
    resized_height = int(height * scaling_factor)

    # Resize the frame or image
    frame_resized = cv2.resize(frame, (resized_width, resized_height))

    # Add the filename as overlay text
    if mode == 'video':
        cv2.putText(frame_resized, filename, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, .5, (255, 255, 255), 1)
        image_resized = Image.fromarray(frame_resized)
    elif mode == 'image':
        image_resized = image.resize((resized_width, resized_height))
        draw = ImageDraw.Draw(image_resized)
        font = ImageFont.load_default()  # Use a default font # Adjust the font size to match the video text size
        draw.text((10, 30), filename, fill=(255, 255, 255), font=font)

    # Convert the PIL image to a format that can be displayed on the Tkinter canvas
    photo = ImageTk.PhotoImage(image=image_resized)

    # Display the image on the canvas
    canvas.create_image(0, 0, image=photo, anchor='nw')
    canvas.image = photo  # Keep a reference!

# Function to crop and save videos with a progress bar
def process_video(rect, progress_label, progress_bar, videos, scaling_factor):
    start, end, label, _ = rect
    x1, y1 = int(start[0] / scaling_factor), int(start[1] / scaling_factor)
    x2, y2 = int(end[0] / scaling_factor), int(end[1] / scaling_factor)
    cropped_video_path = f'{label}.avi'
    vid = cv2.VideoCapture(video_path)
    fourcc = cv2.VideoWriter_fourcc(*'XVID')
    out = cv2.VideoWriter(cropped_video_path, fourcc, vid.get(cv2.CAP_PROP_FPS), (x2 - x1, y2 - y1))

    while True:
        ret, frame = vid.read()
        if not ret:
            break
        crop_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)[y1:y2, x1:x2]
        out.write(cv2.cvtColor(crop_img, cv2.COLOR_RGB2BGR))

    out.release()
    vid.release()
    videos.append(Video(label, cropped_video_path))

    # Update progress in the main thread
    root.after(100, lambda: update_progress(progress_label, progress_bar))

def crop_and_save_videos(callback):
    global scaling_factor, root, video_paths, media_rectangles
    progress = tk.Toplevel(root)
    progress.title("Cropping Videos")

    # Text label to display the progress status
    progress_label = Label(progress, text="Starting...")
    progress_label.pack(pady=(10, 0))

    # Initialize the progress bar
    total_rectangles = sum(len(rects) for rects in media_rectangles.values())
    progress_bar = Progressbar(progress, orient=tk.HORIZONTAL, length=300, mode='determinate', maximum=total_rectangles)
    progress_bar.pack(pady=20)  # Display the progress bar in the window

    videos = []  # This will store Video objects
    count = [0]  # Use a mutable object so the inner function can modify it

    def update_progress():
        # Use count[0] to track progress, increment when a video is processed
        count[0] += 1
        progress_label.config(text=f"Cropping video {count[0]}/{total_rectangles}")
        progress_bar['value'] = count[0]
        root.update_idletasks()

    def process_single_video(video_path, rect, cropped_videos_dir):
        start, end, label, original_file_name = rect
        cropped_video_path = os.path.join(cropped_videos_dir, f'{label}.avi')
        x1, y1 = int(start[0] / scaling_factor), int(start[1] / scaling_factor)
        x2, y2 = int(end[0] / scaling_factor), int(end[1] / scaling_factor)
        vid = cv2.VideoCapture(video_path)
        fourcc = cv2.VideoWriter_fourcc(*'XVID')
        out = cv2.VideoWriter(cropped_video_path, fourcc, vid.get(cv2.CAP_PROP_FPS), (x2 - x1, y2 - y1))

        while True:
            ret, frame = vid.read()
            if not ret:
                break
            crop_img = frame[y1:y2, x1:x2]
            out.write(crop_img)

        out.release()
        vid.release()

        # Save the mapping of well label to original file name in the hidden file
        hidden_file_path = os.path.join(cropped_videos_dir, '.well_original_mapping.txt')
        with open(hidden_file_path, 'a') as hidden_file:
            hidden_file.write(f"{label},{original_file_name}\n")

        videos.append(Video(label, cropped_video_path, original_file_name))
        # Instead of directly calling update_progress, schedule it to be called on the main thread
        root.after(0, update_progress)

    # Create and start threads for each video and its rectangles
    threads = []  # List to keep track of threads
    for video_path in video_paths:
        if video_path in media_rectangles:  # Check if there are rectangles for this video
            for rect in media_rectangles[video_path]:
                thread = threading.Thread(target=process_single_video, args=(video_path, rect, cropped_videos_dir))
                thread.start()
                threads.append(thread)

    # Use a callback loop to wait for all threads to complete
    def check_threads():
        if any(thread.is_alive() for thread in threads):
            # Reschedule check
            root.after(100, check_threads)
        else:
            # Once all threads are done, remove the progress widgets and close the progress window
            progress_label.pack_forget()
            progress_bar.pack_forget()
            progress.destroy()  # Close the progress window
            if callback:
                callback(videos)  # Ensure this function displays the cropped videos

    check_threads()

def add_overlay_text(frame, text, position):
    font = cv2.FONT_HERSHEY_SIMPLEX
    cv2.putText(frame, text, position, font, 1.5, (255, 255, 255), 2, cv2.LINE_AA)

def update_progress(progress_label, progress_bar):
    progress_bar['value'] += 1
    processed_count = progress_bar['value']
    total_count = progress_bar['maximum']
    progress_label.config(text=f"Processing: {processed_count}/{total_count}")


def review_and_classify_videos(videos, cropped_videos_dir, results_path=None):
    global review_complete, stage_names
    control_panel_height = 150  # Height of the control panel at the bottom
    text_color = (255, 255, 255)  # White color for text
    control_bg_color = (0, 0, 0)  # Black background for the control panel
    text_font = cv2.FONT_HERSHEY_SIMPLEX

    for video in videos:
        cap = cv2.VideoCapture(video.cropped_path)
        window_name = 'Review'
        cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
        cv2.resizeWindow(window_name, screen_width, screen_height - control_panel_height)
        video_display_height = screen_height - control_panel_height
        pause = True  # Start with the video paused
        frame_jump = 120  # Small jump
        big_jump = 1440  # Large jump
        playback_speed = 1  # Initial playback speed
        min_playback_speed = 0.25  # Minimum playback speed
        max_playback_speed = 10  # Maximum playback speed
        stage_times = {}

        ret, last_frame = cap.read()  # Read the first frame to display when paused

        while True:
            if not pause:
                ret, frame = cap.read()
                if not ret:
                    cap.set(cv2.CAP_PROP_POS_FRAMES, 0)  # Loop the video
                    continue
            else:
                frame = last_frame  # Display the last frame if paused

            # Resize frame to fit the adjusted display height
            resized_frame = cv2.resize(frame, (screen_width, video_display_height))

            # Create a separate frame for the control panel
            control_panel = np.zeros((control_panel_height, screen_width, 3), dtype=np.uint8)
            control_panel[:] = control_bg_color

            # Display instructions on the control panel
            cv2.putText(control_panel, "Controls: Play/Pause - 'p', Forward - 'f', Backward - 'b',",
                        (10, 60), text_font, 0.7, text_color, 2)
            cv2.putText(control_panel,
                        "Jump Forward - 'g', Jump Backward - 'n', Speed Up - 'u', Speed Down - 'd', Quit - 'q'",
                        (10, 90), text_font, 0.7, text_color, 2)
            cv2.putText(control_panel, f"Speed: {playback_speed}x", (10, 120),
                        text_font, 0.7, text_color, 2)
            # Display current time on video
            current_time = cap.get(cv2.CAP_PROP_POS_MSEC) / 1000
            cv2.putText(control_panel, f"Time: {current_time:.2f}s", (10, 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)

            # Overlay the control panel on top of the resized frame
            combined_frame = np.vstack((resized_frame, control_panel))

            # Show the combined frame
            cv2.imshow(window_name, combined_frame)

            key = cv2.waitKey(int(25 / playback_speed)) & 0xFF
            if key == ord('q'):
                break
            elif key == ord('p'):
                pause = not pause
            elif key == ord('f'):
                frame_index = cap.get(cv2.CAP_PROP_POS_FRAMES) + frame_jump
                cap.set(cv2.CAP_PROP_POS_FRAMES, frame_index)
            elif key == ord('b'):
                frame_index = max(cap.get(cv2.CAP_PROP_POS_FRAMES) - frame_jump, 0)
                cap.set(cv2.CAP_PROP_POS_FRAMES, frame_index)
            elif key == ord('g'):
                frame_index = cap.get(cv2.CAP_PROP_POS_FRAMES) + big_jump
                cap.set(cv2.CAP_PROP_POS_FRAMES, frame_index)
            elif key == ord('n'):
                frame_index = max(cap.get(cv2.CAP_PROP_POS_FRAMES) - big_jump, 0)
                cap.set(cv2.CAP_PROP_POS_FRAMES, frame_index)
            elif key == ord('u') and playback_speed < max_playback_speed:
                playback_speed = min(playback_speed * 2, max_playback_speed)
            elif key == ord('d') and playback_speed > min_playback_speed:
                playback_speed = max(playback_speed / 2, min_playback_speed)

            last_frame = frame  # Save the last frame when paused

        video.set_stage_times(stage_times)
        cap.release()
        cv2.destroyAllWindows()
        # Collecting stage times after video is closed
        collect_stage_times(video, window_name)
        save_classifications_to_excel(video, results_path, video.original_file_name, cropped_videos_dir)

    review_complete = True



# Function to save classifications to an Excel spreadsheet
def save_classifications_to_excel(video, results_path, original_file_name, cropped_videos_dir):
    global stage_names  # Assuming this is a global variable with user-defined stage names
    if not results_path:
        filename = f'Blinded_Analysis_Results.xlsx'
        results_path = os.path.join(cropped_videos_dir, filename)

    if not os.path.exists(results_path):
        wb = Workbook()
        ws = wb.active
        headers = ['Well Label', 'Original File'] + stage_names
        ws.append(headers)
    else:
        wb = load_workbook(results_path)
        ws = wb.active

    row = [video.label, original_file_name] + [video.stage_times.get(stage, 'N/A') for stage in stage_names]
    ws.append(row)
    wb.save(results_path)


def save_classifications_to_excel_pic(image, cropped_images_dir):
    global stage_names
    results_path = os.path.join(cropped_images_dir, "Image_Analysis_Results.xlsx")
    if not os.path.exists(results_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        headers = ["Well Label", "Original File"] + stage_names
        ws.append(headers)
    else:
        wb = load_workbook(results_path)
        ws = wb.active

    row = [image.label, image.original_file_name] + [image.stage_times.get(stage, 'N/A') for stage in stage_names]
    ws.append(row)
    wb.save(results_path)





def save_randomization_order(videos, video_name):
    order_filename = f'.{video_name}_Randomization_Order.txt'
    order_path = os.path.join(cropped_videos_dir, order_filename)

    with open(order_path, 'w') as file:
        for video in videos:
            file.write(f'{video.label}\n')

def resume_analysis(cropped_videos_dir):
    global review_complete, results_path, stage_names
    # Load the randomization order and results file
    video_name = os.path.basename(os.path.normpath(cropped_videos_dir)).replace('_Cropped_Videos', '')
    order_filename = f'.{video_name}_Randomization_Order.txt'
    order_path = os.path.join(cropped_videos_dir, order_filename)

    # Find the .xlsx file in the folder
    results_path = None
    for file in os.listdir(cropped_videos_dir):
        if file.endswith(".xlsx"):
            results_filename = file
            break
    else:
        messagebox.showerror("Error", "No .xlsx file found in the folder.")
        return

    results_path = os.path.join(cropped_videos_dir, results_filename)
    wb = load_workbook(results_path)
    ws = wb.active

    # Infer stage names from the Excel file header
    stage_names = [cell.value for cell in ws[1]][2:]  # Skip 'Well Label' and 'Original File'

    with open(order_path, 'r') as file:
        randomized_order = [line.strip() for line in file.readlines()]

    # Load the well to original file path mapping from the hidden file
    hidden_file_path = os.path.join(cropped_videos_dir, '.well_original_mapping.txt')
    well_to_original = {}
    with open(hidden_file_path, 'r') as hidden_file:
        for line in hidden_file:
            well_label, original_file_path = line.strip().split(',')
            well_to_original[well_label] = original_file_path

    last_analyzed_well = ws.cell(row=ws.max_row, column=1).value

    if last_analyzed_well in randomized_order:
        start_index = randomized_order.index(last_analyzed_well) + 1
        videos_to_analyze = [
            Video(label, os.path.join(cropped_videos_dir, f'{label}.avi'), well_to_original.get(label, 'Unknown'))
            for label in randomized_order[start_index:]
        ]
        review_and_classify_videos(videos_to_analyze, cropped_videos_dir, results_path)
        if review_complete:
            root.quit()
    else:
        messagebox.showerror("Error", "Could not find the last analyzed well in the randomization order.")

# GUI for selecting the video file
def select_video_files():
    global video_paths, stage_names
    video_paths = filedialog.askopenfilenames(
        title="Select video files",
        filetypes=[("MP4 files", "*.mp4"), ("AVI files", "*.avi"), ("All files", "*.*")]
    )
    if video_paths:
        stage_names = get_stage_names()  # Set stage names here
        first_frame = get_first_frame(video_paths[0])
        if first_frame is not None:
            update_canvas(first_frame)
        else:
            messagebox.showerror("Error", "Could not read the first frame of the video.")


# Function to finish selection and start cropping
def finish_selection():
    global video_paths, review_complete, cropped_videos_dir, results_path
    canvas.pack_forget()  # Hide the canvas

    # Create a directory for cropped videos based on the video file name
    cropped_videos_dir = os.path.join(Path(video_paths[0]).parent, "Cropped_Videos")
    os.makedirs(cropped_videos_dir, exist_ok=True)

    def after_cropping(videos):
        # Shuffle the list of videos and save the randomization order inside the Cropped_Videos directory
        random.shuffle(videos)
        save_randomization_order(videos, os.path.basename(cropped_videos_dir))  # Pass the base name of the cropped_videos_dir
        review_and_classify_videos(videos, cropped_videos_dir)  # Pass the cropped_videos_dir directly
        if review_complete:
            root.quit()

    crop_and_save_videos(after_cropping)

import cv2

def image_to_single_frame_video(image_path, output_video_path):
    img = cv2.imread(image_path)
    height, width, layers = img.shape
    video = cv2.VideoWriter(output_video_path, cv2.VideoWriter_fourcc(*'MP4V'), 1, (width, height))
    video.write(img)
    video.release()


def convert_image_to_video(image_path, video_path):
    """Converts an image into a one-frame video."""
    img = cv2.imread(image_path)
    height, width, layers = img.shape
    video = cv2.VideoWriter(video_path, cv2.VideoWriter_fourcc(*'mp4v'), 1, (width, height))
    video.write(img)
    video.release()

def select_image_files():
    global image_paths, stage_names, video_paths, mode
    image_paths = filedialog.askopenfilenames(
        title="Select image files",
        filetypes=[("JPG files", "*.jpg"), ("JPEG files", "*.jpeg"), ("PNG files", "*.png"), ("BMP files", "*.bmp"),
                   ("All files", "*.*")]
    )
    if image_paths:
        stage_names = get_stage_names()
        # Convert images to single-frame videos
        video_paths = []
        for image_path in image_paths:
            video_path = os.path.splitext(image_path)[0] + '.mp4'
            convert_image_to_video(image_path, video_path)
            video_paths.append(video_path)
        # Update the canvas with the first frame (image) for cropping
        update_canvas(image_paths[0])
        # Change mode to 'video' for further processing
        mode = 'video'
    else:
        messagebox.showerror("Error", "No images selected.")




def display_image_on_canvas(image_path):
    global canvas, scaling_factor
    image = Image.open(image_path)
    width, height = image.size

    # Define a margin to account for the window's title bar and other UI elements
    ui_margin = 200  # Adjust this value based on your UI

    # Adjust the available screen height to account for the margin
    available_height = screen_height - ui_margin

    # Calculate the scaling factor based on the adjusted available height and screen width
    scaling_factor = min(screen_width / width, available_height / height, 1)

    resized_width = int(width * scaling_factor)
    resized_height = int(height * scaling_factor)
    image = image.resize((resized_width, resized_height))
    photo = ImageTk.PhotoImage(image)

    # Center the image on the canvas
    x_position = (screen_width - resized_width) // 2
    y_position = (available_height - resized_height) // 2

    canvas.create_image(x_position, y_position, image=photo, anchor=NW)
    canvas.config(scrollregion=canvas.bbox(ALL))
    canvas.image = photo  # Keep a reference!


def crop_and_save_images(callback):
    global scaling_factor, root, image_paths, media_rectangles, cropped_images_dir
    if not image_paths:
        messagebox.showinfo("No Images Selected", "Please select image files to crop.")
        return
    if not media_rectangles:
        messagebox.showinfo("No Selections Made", "Please draw rectangles on the images to specify areas to crop.")
        return

    cropped_images_dir = os.path.join(Path(image_paths[0]).parent, f"{Path(image_paths[0]).stem}_Cropped_Images")
    os.makedirs(cropped_images_dir, exist_ok=True)

    images = []
    for image_path in image_paths:
        if image_path in media_rectangles:
            for rect in media_rectangles[image_path]:
                start, end, label, original_file_name = rect
                # Apply the inverse of the scaling factor to the coordinates
                x1, y1 = int(start[0] / scaling_factor), int(start[1] / scaling_factor)
                x2, y2 = int(end[0] / scaling_factor), int(end[1] / scaling_factor)
                image = Image.open(image_path)
                crop_img = image.crop((x1, y1, x2, y2))
                cropped_image_path = os.path.join(cropped_images_dir, f'{label}.png')
                crop_img.save(cropped_image_path)
                images.append(Video(label, cropped_image_path, original_file_name))

    if callback:
        callback(images, cropped_images_dir)


def custom_ask_question(title, message):
    dialog = Toplevel(root)
    dialog.title(title)
    dialog.grab_set()  # Make the dialog modal

    # Display the message
    Label(dialog, text=message).pack(pady=10)

    # Variable to store the response
    response = StringVar(dialog, value='cancel')

    # Define the callback function for the buttons
    def set_response_and_close(res):
        response.set(res)
        dialog.destroy()

    # Create buttons for the custom options
    Button(dialog, text="Start New Analysis", command=lambda: set_response_and_close('new')).pack(side=LEFT, padx=5)
    Button(dialog, text="Resume Analysis", command=lambda: set_response_and_close('resume')).pack(side=LEFT, padx=5)
    Button(dialog, text="Cancel", command=lambda: set_response_and_close('cancel')).pack(side=LEFT, padx=5)

    dialog.wait_window()  # Wait for the dialog to be closed
    return response.get()



def select_media_file():
    global mode
    response = custom_ask_question("Analysis", "Do you want to start a new analysis or resume an existing one?")
    if response == 'new':
        if mode == 'video':
            select_video_files()
        elif mode == 'image':
            select_image_files()
    elif response == 'resume':
        if mode == 'video':
            directory = filedialog.askdirectory(title="Select directory with cropped videos")
            if directory:
                resume_analysis(directory)
        elif mode == 'image':
            directory = filedialog.askdirectory(title="Select directory with cropped images")
            if directory:
                resume_analysis(directory)
    # No action is needed for the 'cancel' response, as it will simply close the dialog without doing anything



def finish_media_selection():
    global mode
    if mode == 'video':
        finish_selection()  # Correct function name
    elif mode == 'image':
        finish_image_selection()

def set_mode_to_video():
    global mode
    mode = 'video'
    btn_select_media.config(text="Select Video File")
    btn_finish_media.config(text="Finish Video Selection")

def set_mode_to_image():
    global mode
    mode = 'image'
    btn_select_media.config(text="Select Image File")
    btn_finish_media.config(text="Finish Image Selection")

def categorize_image(image, results_path):
    global stage_names
    img = Image.open(image.cropped_path)
    image_window = Toplevel(root)
    image_window.title("Review Image")
    photo = ImageTk.PhotoImage(img)
    label = Label(image_window, image=photo)
    label.image = photo
    label.pack()

    stage_times = {}
    for stage in stage_names:
        time = simpledialog.askstring("Input", f"Enter time for {stage} (in seconds):", parent=image_window)
        stage_times[stage] = time if time else 'N/A'

    image.set_stage_times(stage_times)
    save_classifications_to_excel_pic(image, cropped_images_dir)
    image_window.wait_window()  # Wait for the window to be closed before continuing

def review_and_classify_images(images, cropped_images_dir):
    global review_complete, stage_names
    stage_names = get_stage_names()
    for image in images:
        img = Image.open(image.cropped_path)
        image_window = Toplevel(root)
        image_window.title("Review Image")
        photo = ImageTk.PhotoImage(img)
        label_widget = Label(image_window, image=photo)
        label_widget.image = photo
        label_widget.pack()

        stage_times = {}
        for stage in stage_names:
            time = simpledialog.askstring("Input", f"Enter notes/time for {stage}:", parent=image_window)
            stage_times[stage] = time if time else 'N/A'

        image.set_stage_times(stage_times)
        save_classifications_to_excel_pic(image, cropped_images_dir)
        image_window.wait_window()  # Wait for the window to be closed before continuing

    review_complete = True
    if review_complete:
        root.quit()
def after_cropping_images(images, cropped_images_dir):
    save_randomization_order_images(images, cropped_images_dir)
    review_and_classify_images(images, cropped_images_dir)

# In the finish_image_selection function, pass after_cropping_images as the callback
def finish_image_selection():
    global image_paths, cropped_images_dir
    canvas.pack_forget()
    if image_paths:
        image_path = image_paths[0]  # Define image_path here
        image_name = Path(image_path).stem
        cropped_images_dir = os.path.join(Path(image_path).parent, f"{image_name}_Cropped_Images")
        os.makedirs(cropped_images_dir, exist_ok=True)

        def after_cropping(images, cropped_images_dir):  # Updated to match expected callback signature
            global results_path
            if not images:
                messagebox.showinfo("No Images", "No images were cropped.")
                return
            # Set results_path to the Excel file in the cropped images directory
            for file in os.listdir(cropped_images_dir):
                if file.endswith(".xlsx"):
                    results_path = os.path.join(cropped_images_dir, file)
                    break
            else:
                messagebox.showerror("Error", "No .xlsx file found in the cropped images directory.")
                return
            review_and_classify_images(images, cropped_images_dir)

        crop_and_save_images(after_cropping_images)  # Corrected to use the defined after_cropping function

def resume_analysis_images(cropped_images_dir):
    global review_complete, results_path, stage_names
    image_name = os.path.basename(os.path.normpath(cropped_images_dir)).replace('_Cropped_Images', '')
    order_filename = f'.{os.path.basename(image_name)}_Cropped_Images_Randomization_Order.txt'
    order_path = os.path.join(cropped_images_dir, order_filename)

    results_path = None
    for file in os.listdir(cropped_images_dir):
        if file.endswith(".xlsx"):
            results_path = os.path.join(cropped_images_dir, file)
            break
    if not results_path:
        messagebox.showerror("Error", "No .xlsx file found in the folder.")
        return

    wb = load_workbook(results_path)
    ws = wb.active
    stage_names = [cell.value for cell in ws[1]][2:]  # Skip 'Well Label'

    with open(order_path, 'r') as file:
        randomized_order = [line.strip() for line in file.readlines()]

    hidden_file_path = os.path.join(cropped_images_dir, '.well_original_mapping.txt')
    well_to_original = {}
    with open(hidden_file_path, 'r') as hidden_file:
        for line in hidden_file:
            well_label, original_file_path = line.strip().split(',')
            well_to_original[well_label] = os.path.basename(original_file_path)
    last_analyzed_image = ws.cell(row=ws.max_row, column=1).value
    if last_analyzed_image in randomized_order:
        start_index = randomized_order.index(last_analyzed_image) + 1
        images_to_analyze = [
            Video(label, os.path.join(cropped_images_dir, f'{label}.png'), well_to_original.get(label, 'Unknown'))
            for label in randomized_order[start_index:]
        ]
        review_and_classify_images(images_to_analyze, cropped_images_dir)
        if review_complete:
            root.quit()
    else:
        messagebox.showerror("Error", "Could not find the last analyzed image in the randomization order.")


def save_randomization_order_images(images, image_name):
    order_filename = f'.{os.path.basename(image_name)}_Randomization_Order.txt'
    order_path = os.path.join(cropped_images_dir, order_filename)
    print(f"Saving randomization order to: {order_path}")

    random.shuffle(images)  # Shuffle the images list before saving the order

    with open(order_path, 'w') as file:
        for image in images:
            file.write(f'{image.label}\n')
    print("Randomization order saved.")


current_media_index = 0  # Index to track the current media file

def next_media():
    global current_media_index, mode
    if mode == 'video' and current_media_index < len(video_paths) - 1:
        current_media_index += 1
        update_canvas(get_first_frame(video_paths[current_media_index]))
    elif mode == 'image' and current_media_index < len(image_paths) - 1:
        current_media_index += 1
        display_image_on_canvas(image_paths[current_media_index])

def get_stage_names():
    global stage_names
    if stage_names:  # Check if stage_names is already populated
        return stage_names
    stage_count = simpledialog.askinteger("Input", "Enter the number of stages:", parent=root)
    if not stage_count:
        return []
    for i in range(1, stage_count + 1):
        stage_name = simpledialog.askstring("Input", f"Enter name for Stage {i}:", parent=root)
        if stage_name:
            stage_names.append(stage_name)
    return stage_names


# Button to finish selection and start cropping
# Create a frame to hold the buttons
button_frame = Frame(root)
button_frame.pack(side=BOTTOM, pady=10)  # Add some padding for visual separation

# Create and pack buttons within the frame, using 'side=LEFT' to arrange them horizontally
btn_video_mode = Button(button_frame, text="Video Mode", command=set_mode_to_video)
btn_video_mode.pack(side=LEFT, padx=5)  # Add some padding between buttons

btn_image_mode = Button(button_frame, text="Image Mode", command=set_mode_to_image)
btn_image_mode.pack(side=LEFT, padx=5)

btn_next_media = Button(button_frame, text="Next Media", command=next_media)
btn_next_media.pack(side=LEFT, padx=5)


btn_select_media = Button(button_frame, text="Select Media File", command=select_media_file)
btn_select_media.pack(side=LEFT, padx=5)

btn_finish_media = Button(button_frame, text="Finish Selection", command=finish_media_selection)
btn_finish_media.pack(side=LEFT, padx=5)


# Bind the canvas to the rectangle drawing function
canvas.bind("<Button-1>", draw_rectangle)
canvas.bind("<B1-Motion>", draw_rectangle)
canvas.bind("<ButtonRelease-1>", finalize_rectangle)

# Start the Tkinter event loop
root.mainloop()
